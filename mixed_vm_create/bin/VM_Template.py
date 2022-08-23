# -*- coding=utf-8 -*-
#!/usr/bin/python
####################################################
## Copyright (2022, ) Gemini.Chen
## Author: gemini_chen@163.com
## Date  : 2022/07/22
#####################################################

'''
This is a tool based on openpyxl, get the vm information from existing vm template file.
'''
import string, time
import os, sys, shutil, platform, re
from VM_Request import VM_Request
from VM_Quota import Quota
from openpyxl import load_workbook
from Logs import _banner,_banner_index,_log_success,_log_warn,_log_error,_log_info,_log_adv

reload(sys)
sys.setdefaultencoding("utf-8")

class OpenStack:
    def __init__(self, vm_request):
        self.vendorType = "openstack".upper()
        self.vm = vm_request
        self.current_direct = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))

        # verify ip
    def check_ip_in_pool(self, mask, target_ip):
            ip_flag = True
            min_ip = [int(i) for i in mask.split('-')[0].split('.')]
            max_ip = [int(i) for i in mask.split('-')[1].split('.')]
            ip = [int(i) for i in target_ip.split('.')]
            for index in range(4):
                if (int(ip[index]) < int(min_ip[index])) or (int(ip[index]) > int(max_ip[index])):
                    ip_flag = False
            return ip_flag

    def generate_target_openstack_template(self, content_from_excel):
        _banner_index("Task : Generate Target [OpenStack] VMs File From Existing Template ")

        base_content = content_from_excel

        # judge the provider is exiting or not
        if "OPENSTACK" in self.vm.get_available_server():
            _log_success("Cloud [OpenStack] Provider is Available")
        else:
            _log_error("Cloud [OpenStack] Provider is Not Available, Please Double Confirm")
            sys.exit(0)

        # Load the template
        try:
            template_file_data = open(self.current_direct+"/template/openstack_template.json", "r")
            template_file_setting = string.Template(template_file_data.read())
            # generate init file
            input_from_excel = base_content.get('Detail')

            # get the common data from server
            openstack_template_data = self.vm.get_common_data_from_server()

            # write into real init file (re-pack variable)
            for index in range(len(input_from_excel)):

                # verify project id (excel <--> server)
                _log_info("[OpenStack] Project ID Setting")
                project_id_from_excel = str(input_from_excel[index][0]).strip()

                project_id_from_server = self.vm.get_project_list()

                # verify input is valid or not
                if project_id_from_excel in project_id_from_server.values():
                    openstack_template_data['vm_project_id'] = \
                        [key for key, value in project_id_from_server.items() if value == project_id_from_excel ][0]
                else:
                    _log_warn("Project ID [ " + project_id_from_excel + " ] is not Available")
                    sys.exit(0)

                # zone [1]
                _log_info("Available Zone Start")
                project_zone_from_server = self.vm.get_location()
                # verify zone (excel <--> db) [1]
                project_zone_from_excel = str(input_from_excel[index][1]).strip()

                zone_detail = self.vm.get_resource_related(project_zone_from_excel, "openstack")

                if zone_detail.get('status'):
                    # verify az is selected
                    _log_info("Project Zone [" + project_zone_from_excel + " ]  Available")

                else:
                    _log_error("Project Zone [" + project_zone_from_excel + " ] Not Available")
                    sys.exit(0)

                # vm name
                project_vm_name_from_excel = input_from_excel[index][2]
                if project_vm_name_from_excel.strip() is not "":
                    openstack_template_data['vm_name'] = project_vm_name_from_excel
                else:
                    _log_info("VM name is not invalid")

                # verify template
                vm_flavor_from_excel = str(input_from_excel[index][3]).strip()

                # get vm template from server
                vm_flavor_from_server = self.vm.get_vm_template("openstack")

                if vm_flavor_from_excel in vm_flavor_from_server.keys():
                    openstack_template_data['vm_template_sku_id'] = vm_flavor_from_server.get(vm_flavor_from_excel).get(
                        "sku_id")
                    # vm cpu
                    openstack_template_data['vm_cpu'] = vm_flavor_from_server.get(vm_flavor_from_excel).get("cpu")
                    # # vmware capacity
                    openstack_template_data['vm_mem'] = vm_flavor_from_server.get(vm_flavor_from_excel).get("mem")
                    # categoryId
                    openstack_template_data['vm_categoryId'] = vm_flavor_from_server.get(vm_flavor_from_excel).get(
                        "categoryId")
                else:
                    _log_error("[ OpenStack] Flavor [ "+vm_flavor_from_excel+" ] Not Available")
                    sys.exit(0)
                # disk capacity
                vm_disk_from_excel = str(input_from_excel[index][4]).strip()
                vm_disk_from_server = self.vm.get_vm_disk()
                if vm_disk_from_excel in vm_disk_from_server.keys():
                    _log_success("[OpenStack] DISK Setting : PASS")
                    openstack_template_data['vm_disk_sku_id'] = vm_disk_from_server.get(vm_disk_from_excel).get(
                        'sku_id')
                    openstack_template_data['vm_disk_capacity'] = vm_disk_from_server.get(vm_disk_from_excel).get(
                        'capacity')
                    openstack_template_data['vm_disk_catalogId'] = vm_disk_from_server.get(vm_disk_from_excel).get(
                        'catalogId')
                    openstack_template_data['vm_disk_serviceId'] = vm_disk_from_server.get(vm_disk_from_excel).get(
                        'serviceId')
                    openstack_template_data['vm_disk_categoryId'] = vm_disk_from_server.get(vm_disk_from_excel).get(
                        'categoryId')
                else:
                    _log_error("[ OpenStack] Disk [ "+vm_disk_from_excel+" ] Not Available")
                    sys.exit(0)
                # vm image
                vm_image_type_from_excel = str(input_from_excel[index][5]).strip()
                vm_image_version_from_excel = str(input_from_excel[index][6]).strip()

                # Get OpenStack Image
                vm_image_from_server = self.vm.get_openstack_images()
                # here need do some enhancement to ignore the upper or lower

                # fuzz matching : key in list
                import difflib
                check_flag = difflib.get_close_matches(vm_image_type_from_excel.lower(), vm_image_from_server.keys()
                                                       , 1, cutoff=0.5)

                if (vm_image_type_from_excel.lower() in vm_image_from_server.keys() or len(check_flag) > 0) \
                        and (vm_image_version_from_excel in vm_image_from_server.get(str(check_flag[0])).keys()):

                    _log_info(vm_image_type_from_excel + " - - " + vm_image_version_from_excel + " is available")
                    # self.vm_openstack_stage_status['vm_openstack_image'] = True
                    openstack_template_data['vm_os_category'] = str(check_flag[0])

                    openstack_template_data['vm_os_version'] = vm_image_version_from_excel

                    image_version_id = vm_image_from_server.get(str(check_flag[0])).get(vm_image_version_from_excel)
                    openstack_template_data['vm_os_image_id'] = image_version_id

                else:
                    _log_error("[OpenStack] "+vm_image_type_from_excel+"-"+vm_image_version_from_excel+" Not Available")
                    sys.exit(0)

                # network info from excel
                network_name_from_excel = str(input_from_excel[index][7]).strip()
                subnet_name_from_excel = str(input_from_excel[index][8]).strip()

                # network info from server
                openstack_network_flag = False
                vm_network_from_server = self.vm.get_openstack_network()
                for network_index in range(len(vm_network_from_server)):
                    # network info from server
                    network_name_from_server = vm_network_from_server[network_index].get("network_name")
                    # subnet
                    subnet_name_from_server = vm_network_from_server[network_index].get("subnet_name")
                    #
                    if network_name_from_excel == network_name_from_server and \
                            subnet_name_from_excel == subnet_name_from_server:
                        openstack_template_data['vm_network_name'] = network_name_from_server
                        openstack_template_data['vm_subnet_name'] = subnet_name_from_server

                        network_id_from_server = vm_network_from_server[network_index].get("network_id")
                        subnet_id_from_server = eval(str(vm_network_from_server[network_index].get("subnet_detail"))).get('id')

                        # find network id from src
                        openstack_template_data['network_id'] = network_id_from_server
                        openstack_template_data['subnet_id'] = subnet_id_from_server

                        network_detail = vm_network_from_server[network_index].get("subnet_detail")
                        network_detail_out = str(network_detail).replace('u\'', '\'').replace('\"', '\\\"') \
                            .replace("\'", "\"").replace("\"None\"", "null")

                        # network detail
                        openstack_template_data['network_list_detail'] = network_detail_out

                        _log_info("[OpenStack] Network Setting : PASS")
                        openstack_network_flag = True

                        # verify ip is available or not
                        ip_setting_from_excel = str(input_from_excel[index][9]).strip()
                        # subnet pool
                        subnet_pool_from_server = \
                            vm_network_from_server[network_index].get("subnet_detail").get('ipPools')[2:-2]

                        # ip check
                        p = '((\d{1,2})|([01]\d{2})|(2[0-4]\d)|(25[0-5]))'
                        pattern = '^' + '\.'.join([p] * 4) + '$'

                        if len(ip_setting_from_excel) == 0:
                            # no need setting ip
                            openstack_template_data['vm_fixed_ip'] = ""
                        elif len(ip_setting_from_excel) > 0 and bool(re.match(pattern, ip_setting_from_excel)):
                            ip_in_pool = self.check_ip_in_pool(subnet_pool_from_server, ip_setting_from_excel)
                            if ip_in_pool:
                                _log_info("[OpenStack] IP Setting Start")
                                ip_status = 0
                                if platform.system().lower() == "windows":
                                    # windows
                                    ip_status = os.system("ping -n 2 -w 1 " + ip_setting_from_excel)
                                elif platform.system().lower() == "linux":
                                    # linux
                                    ip_status = os.system("ping -c 2 " + ip_setting_from_excel)
                                if ip_status != 0:
                                    # update fixed ip
                                    openstack_template_data['vm_fixed_ip'] = str(ip_setting_from_excel).strip()
                                    _log_success("[OpenStack] IP Setting: PASS")
                                else:
                                    _log_error("[OpenStack] IP [ " +
                                               ip_setting_from_excel + " ] is using , Please Re-Setting")
                                    sys.exit(0)
                            else:
                                _log_error("[OpenStack] IP Setting: FAILED")
                                _log_warn("["+ip_setting_from_excel+"] is not valid in subnet pool")
                                sys.exit(0)
                        else:
                            _log_error("[OpenStack] IP [ " +
                                       ip_setting_from_excel + " ] is invalid,  Please Double Confirm ")
                            sys.exit(0)

                # openstack network verify
                if not openstack_network_flag:
                    _log_error("[OpenStack] Network Setting Meeting Unexpected Error")
                    _log_warn(
                            "[ " + network_name_from_excel + " ] - [ " + subnet_name_from_excel + " ] is not available")
                    sys.exit(0)

                # SecurityGroup
                # data from excel then verify from server
                security_group_from_excel = str(input_from_excel[index][-1]).strip()
                security_group_flag = False
                for i in range(len(self.vm.get_security_group())):
                    if security_group_from_excel in self.vm.get_security_group()[i].values():
                        openstack_template_data['vm_security_group'] = self.vm.get_security_group()[i].get('groupUuid')
                        security_group_flag = True
                if security_group_flag:
                    _log_success("[OpenStack] Security Group Setting : PASS")
                else:
                    _log_error("[OpenStack] Security Group Setting [ " + security_group_from_excel + " ] is invalid,  Please Double Confirm ")
                    sys.exit(0)

                # vm default password
                vm_password_from_excel = str(input_from_excel[index][-2]).strip()
                # update using the setting
                if len(vm_password_from_excel) > 0 and vm_password_from_excel is not None:
                    vm_default_password = os.popen("java -jar "+self.current_direct+"/lib/util.jar "
                                                   + str(vm_password_from_excel)).read().strip()

                    openstack_template_data['vm_default_password'] = vm_default_password
                else:
                    # not setting use default
                    openstack_template_data['vm_default_password'] = self.vm.vm_default_password
                    # self.vm.get_common_data_from_server.get('vm_default_password')

                # use variable
                update_setting_data = template_file_setting.safe_substitute(openstack_template_data)
                try:
                    # write into new file
                    if not os.path.exists(self.current_direct+"/vm_tasks/openstack"):
                        os.mkdir(self.current_direct+"/vm_tasks/openstack")
                    with open(self.current_direct+"/vm_tasks/openstack/openstack_" + str(index) + ".json", 'w') as fp:
                        # generate new file
                        fp.write(update_setting_data)
                        fp.close()
                        #
                        _log_info("Create File [ openstack_" + str(index) + ".json ] : PASS")
                except IOError:
                    _log_warn("write into new init file meet error")
                finally:
                    # fp.close()
                    pass
        except IOError:
            _log_warn("Cloud [OpenStack] : Open Template File Meet Unexpected Error")

    # create target VMs
    def create_target_vm(self):
        _banner_index("Task : Create VMs Task [ OpenStack ] Start.....")
        base_path = self.current_direct+"/vm_tasks/openstack"
        # OpenStack Section
        # load vm template
        template_list = os.listdir(base_path)
        for file_index in template_list:
            file_from_template = os.path.join(base_path, file_index)
            # get vm status
            out = self.vm.create_target_vms(file_from_template)
            if out.get('status'):
                _log_success("[OpenStack] [ "+file_from_template+" ] Create Action: Success")
                _log_info("Message: " + out.get('message'))
                # mark task status as success
            else:
                _log_error("[OpenStack] [ " + file_from_template + " ] Create Action: Failed")
                _log_warn("Error Message: " + out.get('message'))
                # mark task status as fail
                sys.exit(0)


class Vmware:
    def __init__(self, vm_request):
        self.vendorType = "vmware".upper()
        self.disk_format = {"与原格式相同" :"sameAsSource",
                            "厚置备延迟置零" :"flat",
                            "厚置备快速置零" :"thick",
                            "精简置备" :"thin"}
        self.out = []
        self.vm = vm_request
        self.current_direct = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))

    # create target vmware file
    def generate_target_vmware_template(self, content):
        _banner_index("Task : Generate Target VMs  File From Existing Template : [ Vmware ]")
        #
        content = content

        # judge the provider is exiting or not
        if "VMWARE" in self.vm.get_available_server():
            _log_success("Cloud [ VMware ] Provider is Available")
        else:
            _log_error("Cloud [ VMware ] Provider is Not Available, Please Double Confirm")
            sys.exit(0)

        # load the template
        try:
            # os disk
            vmware_template_file_data = open(self.current_direct+"/template/vmware_template.json", "r")
            # os + data disk
            vmware_template_file_data_disk = open(self.current_direct+"/template/vmware_template_data_disk.json", "r")

            # Here Need Update Later
            vmware_template_file_setting = string.Template(vmware_template_file_data.read())

            # data disk
            vmware_template_file_setting_data_disk = string.Template(vmware_template_file_data_disk.read())

            # generate init file
            vmware_input_from_excel = content.get('Detail')
            vmware_template_data = self.vm.get_common_data_from_server()

            # write into real init file (re-pack variable)
            for index in range(len(vmware_input_from_excel)):

                # verify project id (excel <--> server)
                _log_info("Project ID ["+str(vmware_input_from_excel[index][0])+"] Setting Task : Start")
                project_id_from_excel = str(vmware_input_from_excel[index][0]).strip()
                project_id_from_server = self.vm.get_project_list()

                # verify input is valid or not
                if project_id_from_excel in project_id_from_server.values():
                    vmware_template_data['vm_project_id'] = \
                        [key for key, value in project_id_from_server.items() if value == project_id_from_excel ][0]

                    _log_success("Project ID [ "+vmware_input_from_excel[index][0]+" ] Setting Task : PASS")
                else:
                    _log_error("Project ID [ " + project_id_from_excel + " ] is not Available")
                    sys.exit(0)

                # zone
                _log_info("Available Zone Setting : Start")
                project_zone_from_server = self.vm.get_location()

                # verify zone (excel <--> server)
                project_zone_from_excel = str(vmware_input_from_excel[index][1]).strip()
                # mixed verify
                if project_zone_from_excel in project_zone_from_server.values() \
                        and self.vm.get_resource_related(project_zone_from_excel, "vmware"):
                    # verify az is selected
                    _log_success("Project Zone [" + project_zone_from_excel + " ] [ VMware]  Setting : PASS")
                else:
                    _log_error("Project Zone ["+project_zone_from_excel+" ] [ VMware]  Setting : FAILD")
                    # exit
                    sys.exit(0)

                # vm name
                project_vm_name_from_excel = vmware_input_from_excel[index][2]
                if project_vm_name_from_excel.strip() is not "":
                    vmware_template_data['vm_name'] = str(project_vm_name_from_excel).strip()
                    # vm hostname setting
                    vmware_template_data['vm_host_name'] = str(project_vm_name_from_excel).strip().replace("_","").replace("-","")

                    _log_success("[ "+self.vendorType+" ] VM name [ "+project_vm_name_from_excel +" ] Setting : PASS")
                else:
                    _log_error("[ "+self.vendorType+" ] VM name [ "+project_vm_name_from_excel +" ] Setting : FAILD")

                # verify flavor
                vm_flavor_from_excel = str(vmware_input_from_excel[index][3]).strip()

                # get vm template from server
                vm_flavor_from_server = self.vm.get_vm_template(self.vendorType)

                if vm_flavor_from_excel in vm_flavor_from_server.keys():
                    vmware_template_data['vm_template_sku_id'] = vm_flavor_from_server.get(vm_flavor_from_excel).get("sku_id")
                    # vm cpu
                    vmware_template_data['vm_cpu'] = vm_flavor_from_server.get(vm_flavor_from_excel).get("cpu")
                    # # vmware capacity
                    vmware_template_data['vm_mem'] = vm_flavor_from_server.get(vm_flavor_from_excel).get("mem")
                    # categoryId
                    vmware_template_data['vm_categoryId'] = vm_flavor_from_server.get(vm_flavor_from_excel).get("categoryId")
                else:
                    _log_error("[ VMware] Flavor [ "+vm_flavor_from_excel+" ] Not Available")
                    sys.exit(0)
                # disk format
                vm_disk_format_from_excel = str(vmware_input_from_excel[index][5]).strip()
                vm_disk_format = self.disk_format.get(vm_disk_format_from_excel)

                # update disk format
                vmware_template_data['vmware_disk_format'] = vm_disk_format

                # disk capacity
                vm_os_disk_from_excel = str(vmware_input_from_excel[index][4]).strip()
                vm_disk_from_server = self.vm.get_vm_disk()
                # data disk capacity
                vm_data_disk_from_excel = vmware_input_from_excel[index][6]

                # update os disk
                vmware_template_data['vmware_os_disk_capacity'] = vm_os_disk_from_excel

                if vm_os_disk_from_excel in vm_disk_from_server.keys():
                    vmware_template_data['vm_os_disk_sku_id'] = vm_disk_from_server.get(vm_os_disk_from_excel).get('sku_id')
                    vmware_template_data['vm_os_disk_capacity'] = vm_disk_from_server.get(vm_os_disk_from_excel).get('capacity')
                    vmware_template_data['vm_os_disk_catalogId'] = vm_disk_from_server.get(vm_os_disk_from_excel).get('catalogId')
                    vmware_template_data['vm_os_disk_serviceId'] = vm_disk_from_server.get(vm_os_disk_from_excel).get('serviceId')

                    _log_success("VMware [OS DISK] Setting : PASS")

                # vmware image setting
                vm_image_type_from_excel = str(vmware_input_from_excel[index][7]).strip()
                vm_image_version_from_excel = str(vmware_input_from_excel[index][8]).strip()

                vmware_related = self.vm.get_vmware_related_from_template()

                # here need do some enhancement to ignore the upper or lower
                for related_index in range(len(vmware_related)):
                    # fuzz matching : key in list
                    import difflib
                    type_check_flag = difflib.get_close_matches(vm_image_type_from_excel.lower(),
                                                                vmware_related[related_index].values(), 1, cutoff=0.5)

                    version_check_flag = difflib.get_close_matches(vm_image_version_from_excel.lower(),
                                                                vmware_related[related_index].values(), 1, cutoff=1)
                    if (vm_image_type_from_excel.lower() in vmware_related[related_index].values()
                        or len(type_check_flag) > 0)\
                            and (vm_image_version_from_excel in vmware_related[related_index].values()
                                 or len(version_check_flag) > 0):
                        # vmware image template setting
                        vmware_template_data['vm_os_image_id'] = int(vmware_related[related_index].get('os_image_id'))
                        vmware_template_data['vm_os_category'] = vmware_related[related_index].get('os_type')
                        vmware_template_data['vm_os_version'] = float(vmware_related[related_index].get('os_type_version'))
                        vmware_template_data['os_type_version_card_id'] = vmware_related[related_index].get(
                            'os_type_version_card_id')

                        _log_success("VMware Image Related Setting : PASS")
                    else:
                        _log_error("VMware Image Related Setting : Faild")
                        _log_warn(vm_image_type_from_excel + "-" + vm_image_version_from_excel + " is not available")
                        sys.exit(0)

                # network info from excel
                network_name_from_excel = str(vmware_input_from_excel[index][9]).strip()
                pool_name_from_excel = str(vmware_input_from_excel[index][10]).strip()

                # network info from server
                vm_network_from_server = self.vm.get_vmware_network()

                vmware_network_flag = False
                for network_index in range(len(vm_network_from_server)):
                    pools_detail = []
                    # network name
                    network_name = vm_network_from_server[network_index].get('portGroupName')
                    # pools setting
                    pools_detail.append(vm_network_from_server[network_index].get('pool_detail'))
                    # pool detail
                    pool_name = vm_network_from_server[network_index].get('pool_detail').get('name')

                    if (network_name_from_excel == network_name) and (pool_name_from_excel == pool_name):
                        # portGroupId setting
                        vmware_template_data['poolGroup_id'] = vm_network_from_server[network_index].get('poolGroupId')
                        vmware_template_data['portGroup_id'] = vm_network_from_server[network_index].get('portGroupId')
                        pool_id = vm_network_from_server[network_index].get('pool_detail').get('id')

                        # pool id
                        vmware_template_data['vmware_pool_id'] = pool_id

                        pools_detail_data = str(pools_detail).replace('u\'', '\'').replace('\"', '\\\"') \
                            .replace("\'", "\"").replace("\"None\"", "null")

                        vmware_template_data['pools_details'] = pools_detail_data
                        vmware_network_flag = True

                if vmware_network_flag:
                    _log_success("[Vmware] Network Setting PASS")
                else:
                    _log_error("[Vmware] Network Setting Meeting Unexpected Error")
                    _log_warn(
                            "[ " + network_name_from_excel + " ] - [ " + pool_name_from_excel + " ] is not available")
                    sys.exit(0)

                # ip setting
                ip_setting_detail = str(vmware_input_from_excel[index][11]).strip()
                # ip check
                p = '((\d{1,2})|([01]\d{2})|(2[0-4]\d)|(25[0-5]))'
                pattern = '^' + '\.'.join([p] * 4) + '$'
                if len(ip_setting_detail) > 0 and bool(re.match(pattern, ip_setting_detail)):
                    _log_info("Vmware IP Setting Start")

                    if platform.system().lower() == "windows":
                        # windows
                        ip_status = os.system("ping -n 2 -w 1 " + ip_setting_detail)
                    elif platform.system().lower() == "linux":
                        # linux
                        ip_status = os.system("ping -c 2 " + ip_setting_detail)
                    # ping and check with api
                    pool_id = vmware_template_data.get('vmware_pool_id')

                    if ip_status != 0 and self.vm.vmware_ip_verify(pool_id,ip_setting_detail):
                        ip_policy = "Manual"
                        # update
                        vmware_template_data['vmware_ipPolicy'] = ip_policy
                        vmware_template_data['vmware_ip_address'] = str([ip_setting_detail]).replace("\'","\"")
                        _log_success("[VMware] IP Setting: PASS")
                    else:
                        _log_error("[Vmware] IP [ " + ip_setting_detail + " ] is using , Please Re-Setting")
                        sys.exit(0)
                else:
                    _log_warn("[Vmware] IP [ " + ip_setting_detail + " ] is invalid, Using DHCP IP ..... ")
                    ip_policy = "Auto"
                    # update
                    vmware_template_data['vmware_ipPolicy'] = ip_policy
                    vmware_template_data['vmware_ip_address'] = str([])

                # vm password setting
                vm_password_from_excel = str(vmware_input_from_excel[index][-1]).strip()
                if vm_password_from_excel is not "":
                    vm_default_password = os.popen("java -jar "+self.current_direct+"/lib/util.jar "
                                                   + str(vm_password_from_excel)).read().strip()

                    vmware_template_data['vm_default_password'] = vm_default_password
                else:
                    # vm default password
                    vmware_template_data['vm_default_password'] = self.vm.vm_default_password

                # use variable
                # data disk setting
                if vm_data_disk_from_excel is None or str(vm_data_disk_from_excel).strip() is "" \
                            or int(str(vm_data_disk_from_excel).strip()) == 0:
                    _log_info("Data Disk Not Need Setting, Skipping ...")
                    # exclude the data disk
                    update_setting_data = vmware_template_file_setting.safe_substitute(vmware_template_data)

                elif int(str(vm_data_disk_from_excel).strip()) > 0:
                    _log_warn("Data Disk Setting Start...")
                    # pack the data disk
                    vmware_template_data['vm_data_disk_capacity'] = str(vm_data_disk_from_excel).strip()
                    update_setting_data = vmware_template_file_setting_data_disk.safe_substitute(vmware_template_data)
                else:
                    pass

                try:
                    # write into new file
                    if not os.path.exists(self.current_direct+"/vm_tasks/vmware"):
                        os.mkdir(self.current_direct+"/vm_tasks/vmware")
                    with open(self.current_direct+"/vm_tasks/vmware/vmware_"+str(index)+".json", 'w') as fp:
                        # generate new file
                        fp.write(update_setting_data)
                        fp.close()
                        #
                        _log_info("Create File [ vmware_"+str(index)+".json ] : PASS")
                except IOError:
                    _log_error("[ Vmware ] write into new init file meet error: FAILD")
                finally:
                    # fp.close()
                    pass
        except IOError:
            _log_error("[ Vmware ] Open Template File Meet Unexpected Error")

    # create target VMs
    def create_target_vm(self):
        _banner_index("Task : Create VMs Task [ Vmware ] Vms Start.....")
        base_path = self.current_direct+"/vm_tasks/vmware"
        # OpenStack Section
        # load vm template
        template_list = os.listdir(base_path)
        for file_index in template_list:
            file_from_template = os.path.join(base_path, file_index)
            # get vm status
            out = self.vm.create_target_vms(file_from_template)

            if out.get('status'):
                _log_success("[Vmware] [ "+file_from_template+" ] Create Action: Success")
                # mark task status as success
            else:
                _log_error("[Vmware] [ " + file_from_template + " ] Create Action: Faild")
                _log_warn("Error Message: "+out.get('message'))
                # mark task status as fail
                sys.exit(0)

            # sleep 5s
            time.sleep(5)

class Compute:
    def __init__(self):
        self.vendorType = "compute".upper()
        pass


class VM_Object:
    def __init__(self, file_name):
        self.vm_openstack_stage_status = {}
        #
        self.vm = VM_Request()
        config = self.vm.read_config_file()
        self.out = []
        self.vms_list = []
        self.file_name = file_name
        self.current_direct = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
        self.vm_types = config['VM_Type']['vm_types'].encode('unicode-escape').decode('string_escape').strip()

    # get information from excel
    def get_vms_info(self, file_name, sheet_name):
        content = {}
        vm_details = []
        vm_name_list = []
        # open file
        work_book = load_workbook(filename=file_name)
        try:
            # get sheet <sheet_name>
            sheet = work_book[sheet_name]

            # get the max edge
            sheet_max_rows = sheet.max_row
            sheet_max_columns = sheet.max_column

            # look up all data
            cpu_total = 0
            mem_total = 0
            disk_total = 0
            for row in range(2, sheet_max_rows + 1):
                row_content_list = []
                for column in range(1, sheet_max_columns + 1):

                    row_content_list.append(sheet.cell(row, column).value)
                    if column == 3:
                        # name list
                        vm_name_list.append(str(sheet.cell(row, column).value).strip())
                    # flavor
                    if column == 4:
                        # cpu
                        cpu_total += int(str(sheet.cell(row, column).value).strip()[0:1])
                        # mem
                        mem_total += int(str(sheet.cell(row, column).value).strip()[2:3])
                    if column == 5:
                        # disk
                        disk_total += int(str(sheet.cell(row, column).value).strip())
                vm_details.append(row_content_list)
            # re-pack
            content['cpu_total'] = cpu_total
            content['mem_total'] = mem_total
            content['disk_total'] = disk_total
            content["Detail"] = vm_details
            content["vms"] = vm_name_list

        except IOError:
            _log_error("[ "+sheet_name+" ] Not Available, Please Confirm Again.")
            sys.exit(1)
        return content

    # common setting
    def get_common_setting(self, content, vendor_type):
        quota_checking_flag = False
        _banner_index("Task : Get Common Quota Settings ")
        quota = Quota()
        vms_related = quota.get_tenant_vm_quota()
        service_related = quota.get_tenant_service_quota()
        # from excel
        disk_from_excel = int(content.get('disk_total'))
        cpu_from_excel = int(content.get('cpu_total'))
        mem_from_excel = int(content.get('mem_total'))
        # service amount
        vms_count_from_excel = len(content.get('Detail'))
        # from server
        cpu_from_server = int(vms_related.get('meta.cpu'))
        mem_from_server = int(vms_related.get('meta.mem'))
        disk_from_server = int(vms_related.get('meta.disk'))

        if cpu_from_excel <= cpu_from_server and mem_from_excel <= mem_from_server \
                and disk_from_excel <= disk_from_server:
            if vendor_type == "OpenStack":
                openstack_service = service_related.get('openstack.standard.server')
                # ------------
                openstack_service_vpc = service_related.get('openstack.standard.vpc')
                if vms_count_from_excel <= openstack_service:
                    _log_success("[OpenStack] Quota Verifying Result: PASS")
                    quota_checking_flag = True
                else:
                    _log_error("[OpenStack] Quota-Service  Verifying : FAIL")
                    _log_warn("Expect: [" + str(vms_count_from_excel) + "] \t Existing: [" + str(openstack_service)+"]")

            elif vendor_type == "VMware":
                vmware_service = service_related.get('vmware.standard.server')
                if vms_count_from_excel <= vmware_service:
                    _log_success("[VMware] Quota Verifying Result: PASS")
                    quota_checking_flag = True
                else:
                    _log_error("[VMware] Quota-Service Verifying : FAILED")
                    _log_warn("Expect: [" + str(vms_count_from_excel) + "] \t Existing: [" + str(vmware_service)+"]")
        else:
            _log_error("["+vendor_type+"] Quota is not Enough , Please Contact The System Administrator")
            if cpu_from_excel > cpu_from_server:
                _log_warn("Quota Type: CPU")
            elif mem_from_excel > mem_from_server:
                _log_warn("Quota Type: Memory")
            elif disk_from_excel > disk_from_server:
                _log_warn("Quota Type: Disk")
            sys.exit(0)
        return quota_checking_flag

    # prepare for task (init)
    def prepare_for_task(self):
        _log_info("Prepare For Task [Init And Prepare]")
        if not os.path.exists(self.current_direct+"/vm_tasks_history"):
            os.mkdir(self.current_direct+"/vm_tasks_history")

        if not os.path.exists(self.current_direct+"/vm_tasks"):
            os.mkdir(self.current_direct+"/vm_tasks")
        else:
            # check dir is null or not
            if len(os.listdir(self.current_direct+"/vm_tasks")) == 0:
                _log_info("target workspace is ok for use")
            else:
                # check file is null or not
                shutil.rmtree(self.current_direct+"/vm_tasks")
                # os.rmdir("vm_tasks")
                os.mkdir(self.current_direct+"/vm_tasks")

        if "openstack" in str.lower(self.vm_types):
            _banner_index("OpenStack Stage Start....")
            try:
                # get information
                self.out = self.get_vms_info(self.file_name, sheet_name="OpenStack")
                vms_list = self.out.get('vms')

                # common setting(quota)
                common_quota_verify = self.get_common_setting(self.out, vendor_type="OpenStack")
                if common_quota_verify:
                    # Import OpenStack Resource(init)
                    openstack = OpenStack(self.vm)

                    # task generate-verify-0/1
                    openstack.generate_target_openstack_template(self.out)

                    # 请求执行
                    openstack.create_target_vm()

                    # verify
                    self.vm.get_result_from_server("OpenStack", vms_list)
                else:
                    _log_error("[OpenStack] Quota Verifying Status: FAILED")
                    sys.exit(0)

            except ImportError:
                _log_error("Configure Template File Meet Unexpected Error ( ->:Sheet[OpenStack] ).")
                sys.exit(0)

        if "vmware" in str.lower(self.vm_types):
            _banner_index("Vmware Stage Start....")
            try:
                # 获取信息
                self.out = self.get_vms_info(self.file_name, sheet_name="VMware")
                # vms name
                vms_list = self.out.get('vms')
                # common setting(quota)
                common_quota_verify = self.get_common_setting(self.out, vendor_type="VMware")
                if common_quota_verify:
                    # import Vmware Resource(init)
                    vmware = Vmware(self.vm)
                    # 模板任务生成
                    vmware.generate_target_vmware_template(self.out)

                    # 请求执行
                    vmware.create_target_vm()

                    # verify
                    self.vm.get_result_from_server("VMware", vms_list)
                else:
                    _log_error("[Vmware] Quota Verifying Status: FAIL")
                    sys.exit(0)
            except IOError:
                _log_error("Configure Template File Meet Unexpected Error ( ->:Sheet[VMware] ).")
                sys.exit(0)

        if "compute" in str.lower(self.vm_types):
            _banner_index("Compute Stage Start....")
            try:
                # 获取信息
                self.out = self.get_vms_info(self.file_name, sheet_name="Compute")
                # 模板任务生成

                # 请求执行
            except ImportError:
                _log_error("Configure Template File Meet Unexpected Error ( ->:Sheet[VMware] ).")
                sys.exit()

    # create the testbed
    def clean(self):
        _banner_index("Task: Clean The TestBed Start ")
        # backup task history
        import shutil
        current_time = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
        shutil.copy(self.current_direct+"/"+self.file_name, self.current_direct+"/vm_tasks_history/"+current_time+".xlsx")
        # logout
        flag = self.vm.log_out()
        return flag

    # main action
    def run(self):
        # verify create vms status
        _banner("Prepare For Target VMs")
        self.prepare_for_task()

        # clean testbed
        _banner("Clean The Create VMs Tasks")
        if self.clean():
            _log_success("Clean Task PASS")
        else:
            _log_error("Clean Task FAIL")


if __name__ == '__main__':
    pass